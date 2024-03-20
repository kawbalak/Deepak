import ManageData as md #loads and saves data
#from TESTCASE.OptinalWeightsGA import Returns_AP
import TransformData as td #Cleans and transforms data
import FactorTester as ft #Test factors for suitability in models
import numpy as np
#from IPython.core.magic_arguments import runfile

def LoadData(CapRange):
#CapRange can be All, High, Mid, or Low
    SaveFile = 'Data' + CapRange + 'Cap' #Save loaded data in this file
    
    Filename = SaveFile + '.xlsm' #load data in this file
    R1 = md.removeloadeddata(SaveFile+'.pkl') #deletes any previously saved
    D = md.initialize() #Creats a Python dictionary to store data
    D = md.loaddataFormat1(D,Filename)#Load data from spreadsheet
    Success = md.saveproject(D,SaveFile+'.pkl') #Saves the loaded data to a pickel file
    D = md.loadproject(SaveFile+'.pkl')
    return D
def Run_Lab_1(D,SaveFile):
# D could by Dall, Dhigh, Dmid, or Dlow
#SaveFile could be RunNumber_5
    TD = {} #Dictionary for transformed data
    FA = {} #Dictionary for factors
    N = np.int_(5) #Used to control data cleaning, see notes
    #Next 4 lines were added so forward returns will be used
    Returns_AP = D['Returns_AP'].copy()
    TD['ForwardReturns_AP'] = D['ForwardReturns_AP'].copy()
    D.Keys()
    VMalpha_AP = 1.0*FA['Price_Book_AP']['SignalTile_AP'] + 1.0*FA['STMomentum_AP']['SignalTile_AP'] 
    TD['VMalpha_AP'] = VMalpha_AP.copy()
    TransformList = ['ZSCSBS']
    TD['VMalpha_AP'] = td.Transform(TD['VMalpha_AP'],TransformList,D)
    NumTile = 5 #Number of quantiles to use in AnalyzeFactor
    FA['VMalpha_AP'] = ft.AnalyzeFactor('VMalpha',TD['VMalpha_AP'],TD['ForwardReturns_AP'],NumTile)
    NumPointsS2 = 100 #Number of grid points to search, 100 is usually enough
    S2 = ft.Optimize_IRitp_2(TD['Price_Book_AP'],TD['STMomentum_AP'] ,TD['ForwardReturns_AP'],NumPointsS2)
    NumPointsS3 = 100 #Number of grid points to search, 100 is usually enough
    VMalpha_AP = 1.0*FA['Price_Book_AP']['SignalTile_AP'] + 1.0*FA['STMomentum_AP']['SignalTile_AP'] 
    TD['VMalpha_AP'] = VMalpha_AP.copy()
    TransformList = ['ZSCSBS']
    TD['VMalpha_AP'] = td.Transform(TD['VMalpha_AP'],TransformList,D)
    NumTile = 5 #Number of quantiles to use in AnalyzeFactor
    FA['VMalpha_AP'] = ft.AnalyzeFactor('VMalpha',TD['VMalpha_AP'],TD['ForwardReturns_AP'],NumTile)
    NumPointsS2 = 100 #Number of grid points to search, 100 is usually enough
    S2 = ft.Optimize_IRitp_2(TD['Price_Book_AP'],TD['STMomentum_AP'] ,TD['ForwardReturns_AP'],NumPointsS2)
    FA['VMalpha_AP'] = ft.AnalyzeFactor('VMalpha',TD['VMalpha_AP'],TD['ForwardReturns_AP'],NumTile)
    FA['VMalpha_AP']['ICitp']
    #calculate ICITP , combine two variable to get higher

    N = np.int_(5)
    TD['STMomentum_AP'] = D['STMomentum_AP'].copy()
    TD['STMomentum_AP'] = td.CleanData(TD['STMomentum_AP'],N) #Moderate outliers so model building is better
    TransformList = ['ZSBT','ZSCSBS'] #z-score by time, decile crosssection by sector
    TD['STMomentum_AP'] = td.Transform(TD['STMomentum_AP'],TransformList,D)
    NumTile = np.int_(5) #Number of quantiles to use in AnalyzeFactor
    FA['STMomentum_AP'] = ft.AnalyzeFactor('STMomentum',TD['STMomentum_AP'],TD['ForwardReturns_AP'],NumTile)


    N = np.int_(5)
    TD['SurpriseMomentum_AP'] = D['SurpriseMomentum_AP'].copy()
    TD['SurpriseMomentum_AP'] = td.CleanData(TD['SurpriseMomentum_AP'] ,N) #Moderate outliers so model building is better
    TransformList = ['ZSBT','ZSCSBS'] #z-score by time, decile crosssection by sector
    NumTile = 5 #Number of quantiles to use in AnalyzeFactor
    TD['SurpriseMomentum_AP'] = td.Transform(TD['SurpriseMomentum_AP'] ,TransformList,D)
    FA['SurpriseMomentum_AP'] = ft.AnalyzeFactor('SurpriseMomentum_AP',TD['SurpriseMomentum_AP'] ,TD['ForwardReturns_AP'],NumTile)

    N = 5
    TD['Month12ChangeF12MEarningsEstimate_AP'] = D['Month12ChangeF12MEarningsEstimate_AP'].copy()
    TD['Month12ChangeF12MEarningsEstimate_AP'] = td.CleanData(TD['Month12ChangeF12MEarningsEstimate_AP'],N) #Moderate outliers so model building is better
    TransformList = ['ZSBT','ZSCSBS'] #z-score by time, decile crosssection by sector
    NumTile = np.int_(5) #Number of quantiles to use in AnalyzeFactor
    TD['Month12ChangeF12MEarningsEstimate_AP'] = td.Transform(TD['Month12ChangeF12MEarningsEstimate_AP'],TransformList,D)
    FA['Month12ChangeF12MEarningsEstimate_AP'] = ft.AnalyzeFactor('Month12ChangeF12MEarningsEstimate_AP',TD['Month12ChangeF12MEarningsEstimate_AP'] ,TD['ForwardReturns_AP'],NumTile)

    VMalpha_AP = 1.0*FA['Price_Book_AP']['SignalTile_AP'] + 1.0*FA['STMomentum_AP']['SignalTile_AP'] 
    TD['VMalpha_AP'] = VMalpha_AP.copy()
    TransformList = ['ZSCSBS']
    TD['VMalpha_AP'] = td.Transform(TD['VMalpha_AP'],TransformList,D)
    NumTile = 5 #Number of quantiles to use in AnalyzeFactor
    FA['VMalpha_AP'] = ft.AnalyzeFactor('VMalpha',TD['VMalpha_AP'],TD['ForwardReturns_AP'],NumTile)
    NumTile = np.int_(5)
    FA['Price_Book_AP'] = ft.AnalyzeFactor('Price_Book',TD['Price_Book_AP'],TD['ForwardReturns_AP'],NumTile)
    FA['Price_Book_AP'] ['ICitp']


    FA['Price_Book_AP'] 
    ['ICitp']
    
runfile('/Users/kriti/Downloads/Deepak/Midterm2/ManageData.py', wdir='/Users/')
Dall=LoadData('All')
runfile('/Users/kriti/Downloads/Deepak/Midterm2//ManageData.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2/')
Dall=LoadData('All')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2/')
Dall= LoadData ('ALL')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2/')
Dall= LoadData ('ALL')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
Dall=LoadData('All')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/ManageData.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
Dall = LoadData('All')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/ManageData.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
Dall = LoadData('All')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
Dall = LoadData('All')
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')

###########################################################################################
import ManageData as md #loads and saves data
import TransformData as td #Cleans and transforms data
import FactorTester as ft #Test factors for suitability in models
import numpy as np

def LoadData(CapRange):
#CapRange can be All, High, Mid, or Low
    SaveFile = 'Data' + CapRange + 'Cap' #Save loaded data in this file
    Filename = SaveFile + '.xlsm' #load data in this file
    R1 = md.removeloadeddata(SaveFile+'.pkl') #deletes any previously saved
    D = md.initialize() #Creats a Python dictionary to store data
    D = md.loaddataFormat1(D,Filename)#Load data from spreadsheet
    Success = md.saveproject(D,SaveFile+'.pkl') #Saves the loaded data to a pickel file
    D = md.loadproject(SaveFile+'.pkl')
    return D
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
##############################################################################################################
## ---(Mon Nov 20 16:23:19 2023)---
import pickle as pk
import glob
import os
import numpy as np
from scipy import stats
#from pandas.stats.api import ols
from openpyxl import load_workbook


#if (__name__ == "__main__"): # Execute when invoked from command line

def removeloadeddata(SaveFile):
# print('Deleting *.dat files from current directory ...')
# FileList = glob.glob('*.dat')
# for file in FileList:
# os.remove(file)

    print('Deleting ' + SaveFile + ' file from current directory ...')
    FileList = glob.glob(SaveFile)
    for file in FileList:
        os.remove(file)

    return 0

def initialize(): 
    print('Initializing data dictionary D ...')
    D = {}
    return D

def loaddataFormat1(D,Filename):

    print('Loading spreadsheet ' + Filename + ' to dictionary D ...')
    wb = load_workbook(Filename) 
    asset = wb.get_sheet_by_name('asset') 
    D['Asset'] = list()
    D['Sector_Asset'] = list()
    D['Industry_Asset'] = list()
    D['NumAsset'] = 0
    Var = 'FrogsHopping'
    while Var is not None:
        D['NumAsset'] = D['NumAsset'] + 1
        Var = asset.cell(row = D['NumAsset']+1, column = 1).value
        SectorCode = asset.cell(row = D['NumAsset']+1, column = 2).value
        IndustryCode = asset.cell(row = D['NumAsset']+1, column = 3).value
        if Var is not None: 
            D['Asset'].append(Var)
            D['Sector_Asset'].append(SectorCode)
            D['Industry_Asset'].append(IndustryCode)
            D['NumAsset'] = D['NumAsset']-1 

    sector = wb.get_sheet_by_name('sector') 
    D['Sector'] = list()
    D['SectorCode'] = list()
    D['NumSector'] = 0
    Var = 'FrogsHopping'
    while Var is not None:
        D['NumSector'] = D['NumSector'] + 1
        Var = sector.cell(row = D['NumSector']+1, column = 1).value
        SectorCode = sector.cell(row = D['NumSector']+1, column = 2).value
        if Var is not None: 
            D['Sector'].append(Var)
            D['SectorCode'].append(SectorCode)
            D['NumSector'] = D['NumSector']-1 

    industry = wb.get_sheet_by_name('industry') 
    D['Industry'] = list()
    D['IndustryCode'] = list()
    D['NumIndustry'] = 0
    Var = 'FrogsHopping'
    while Var is not None:
        D['NumIndustry'] = D['NumIndustry'] + 1
        Var = industry.cell(row = D['NumIndustry']+1, column = 1).value
        IndustryCode = industry.cell(row = D['NumIndustry']+1, column = 2).value
        if Var is not None: 
            D['Industry'].append(Var)
            D['IndustryCode'].append(IndustryCode)
            D['NumIndustry'] = D['NumIndustry']-1

    data = wb.get_sheet_by_name('data') 
    D['Variable'] = set()
    Var = 'FrogsHopping'
    D['dataNumRows'] = 1
    VarColNum = 1
    while Var is not None:
        D['dataNumRows'] = D['dataNumRows'] + 1
        Var = data.cell(row = D['dataNumRows'], column = VarColNum).value
        if Var is not None: 
            D['Variable'].add(Var)
            D['dataNumRows'] = D['dataNumRows']-1
            D['NumVariable'] = len(D['Variable']) 

    D['Period'] = list()
    D['NumPeriod'] = 0
    Var = 'FrogsHopping'
    while Var is not None:
        D['NumPeriod'] = D['NumPeriod'] + 1
        Var = data.cell(row = 1, column = D['NumPeriod']+2).value
        if Var is not None: 
            D['Period'].append(Var)
            D['NumPeriod'] = D['NumPeriod']-1 

    NumAsset = D['NumAsset']
    NumPeriod = D['NumPeriod']
    NAN_AP = np.nan*np.zeros((NumAsset,NumPeriod))
    for Var in D['Variable']:
        exec('D[' + "'" + Var + '_AP' + "'" +']' + ' = NAN_AP.copy()')
    # exec('D.' + Var + '_AP = np.nan*np.zeros((D['NumAsset'], D['NumPeriod']))')
        print('Loading variables by asset and period from sheet data ...') 
        Var = 'FrogsHopping'
    dataNumRows = 1
    varColNum = 1
    assetColNum = 2
    while Var is not None:
        dataNumRows = dataNumRows + 1
        Var = data.cell(row = dataNumRows, column = varColNum).value
        asset = data.cell(row = dataNumRows, column = assetColNum).value
    # print(Var)
    # print([dataNumRows,varColNum,assetColNum])
    if Var is not None:
        assetpos = D['Asset'].index(asset)
    for period in D['Period']:
        periodpos = D['Period'].index(period)
        Number = data.cell(row = dataNumRows, column = periodpos+3).value
    # print([assetpos,periodpos,Number])
    # print('D[' + "'" + Var + '_AP' + "'" +']' + '[assetpos,periodpos] = Number')
    exec('D[' + "'" + Var + '_AP' + "'" +']' + '[assetpos,periodpos] = Number')
    # print('D.'+Var+'_AP[assetpos,periodpos] = Number')
    print('Variables by asset and period from sheet data are loaded.') 
    # print(D.AnalystAgreementRevisions_AP[1,1])
    print('Number of rows in sheet data: ' + str(D['dataNumRows']))
    print(' Number of Assets: ' + str(D['NumAsset']))
    print(' Number of Periods: ' + str(D['NumPeriod']))
    print(' Number of Sectors: ' + str(D['NumSector']))
    print(' Number of Industrys: ' + str(D['NumIndustry']))
    print(' Number of Variables: ' + str(D['NumVariable']))
    ForwardReturns_AP = NAN_AP.copy()
    Returns_AP = NAN_AP.copy()
    for p in range(1,D['NumPeriod']):
        for a in range(0,D['NumAsset']):
            Returns_AP[a,p] = np.log(D['Price_AP'][a,p]/D['Price_AP'][a,p-1])
    for p in range(0,D['NumPeriod']-1):
        for a in range(0,D['NumAsset']):
            ForwardReturns_AP[a,p] = np.log(D['Price_AP'][a,p+1]/D['Price_AP'][a,p]) 
            D['ForwardReturns_AP'] = ForwardReturns_AP
            D['Returns_AP'] = Returns_AP
            return D

def saveproject(D,SaveFile):
    print('Saving D to ' + SaveFile)
    handle = open(SaveFile, 'wb')
    pk.dump(D,handle)
    print('D has been saved to ' + SaveFile)
    return 1

def loadproject(SaveFile):
    print('Loading from ' + SaveFile)
    file = open(SaveFile, 'rb')
    D = pk.load(file)
    print('Data loaded from ' + SaveFile)
    return D


def savefile(SaveFile,X):
    print('Saving to ' + SaveFile)
    handle = open(SaveFile, 'wb')
    pk.dump(X,handle)
    print('saved to ' + SaveFile)
    return 1

def loadsavedfile(SaveFile):
    print('Loading from ' + SaveFile)
    file = open(SaveFile, 'rb')
    F = pk.load(file)
    print('Loaded from ' + SaveFile)
    return F
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
#############################################################################################################################################
import ManageData as md #loads and saves data
import TransformData as td #Cleans and transforms data
import FactorTester as ft #Test factors for suitability in models
import numpy as np

def LoadData(CapRange):
#CapRange can be All, High, Mid, or Low
    SaveFile = 'Data' + CapRange + 'Cap' #Save loaded data in this file
    Filename = SaveFile + '.xlsm' #load data in this file
    R1 = md.removeloadeddata(SaveFile+'.pkl') #deletes any previously saved
    D = md.initialize() #Creats a Python dictionary to store data
    D = md.loaddataFormat1(D,Filename)#Load data from spreadsheet
    Success = md.saveproject(D,SaveFile+'.pkl') #Saves the loaded data to a pickel file
    D = md.loadproject(SaveFile+'.pkl')
    return D
    Dall=LoadData('All')
def Run_Lab_1(D,SaveFile):
# D could by Dall, Dhigh, Dmid, or Dlow
#SaveFile could be RunNumber_5
    TD = {} #Dictionary for transformed data
    FA = {} #Dictionary for factors
    N = np.int_(5) #Used to control data cleaning, see notes
    #Next 4 lines were added so forward returns will be used
    Returns_AP = D['Returns_AP'].copy()
    TD['ForwardReturns_AP'] = D['ForwardReturns_AP'].copy()
    D=Dall
    TD = {} #Dictionary for transformed data
    FA = {} #Dictionary for factors
    N = np.int_(5) #Used to control data cleaning, see notes
    #Next 4 lines were added so forward returns will be used
    Returns_AP = D['Returns_AP'].copy()
    TD['ForwardReturns_AP'] = D['ForwardReturns_AP'].copy()

    TD['Price_Book_AP'] = -D['Price_Book_AP'].copy() #Change the sign of price to book because high is bad, low is good
    TD['Price_Book_AP'] = td.CleanData(TD['Price_Book_AP'],N) #Moderate outliers so model building is better
    TransformList = ['ZSBT','ZSCSBS'] #z-score by time, decile crosssection by sector
    TD['Price_Book_AP'] = td.Transform(TD['Price_Book_AP'],TransformList,D)
    X = TD['Price_Book_AP']

    NumTile = np.int_(5)
    FA['Price_Book_AP'] = ft.AnalyzeFactor('Price_Book',TD['Price_Book_AP'],TD['ForwardReturns_AP'],NumTile)

    N = np.int_(5)
    TD['STMomentum_AP'] = D['STMomentum_AP'].copy()
    TD['STMomentum_AP'] = td.CleanData(TD['STMomentum_AP'],N) #Moderate outliers so model building is better
    TransformList = ['ZSBT','ZSCSBS'] #z-score by time, decile crosssection by sector
    TD['STMomentum_AP'] = td.Transform(TD['STMomentum_AP'],TransformList,D)
    NumTile = np.int_(5) #Number of quantiles to use in AnalyzeFactor
    FA['STMomentum_AP'] = ft.AnalyzeFactor('STMomentum',TD['STMomentum_AP'],TD['ForwardReturns_AP'],NumTile)


    N = np.int_(5)
    TD['SurpriseMomentum_AP'] = D['SurpriseMomentum_AP'].copy()
    TD['SurpriseMomentum_AP'] = td.CleanData(TD['SurpriseMomentum_AP'] ,N) #Moderate outliers so model building is better
    TransformList = ['ZSBT','ZSCSBS'] #z-score by time, decile crosssection by sector
    NumTile = 5 #Number of quantiles to use in AnalyzeFactor
    TD['SurpriseMomentum_AP'] = td.Transform(TD['SurpriseMomentum_AP'] ,TransformList,D)
    FA['SurpriseMomentum_AP'] = ft.AnalyzeFactor('SurpriseMomentum_AP',TD['SurpriseMomentum_AP'] ,TD['ForwardReturns_AP'],NumTile)

    N = 5
    TD['Month12ChangeF12MEarningsEstimate_AP'] = D['Month12ChangeF12MEarningsEstimate_AP'].copy()
    TD['Month12ChangeF12MEarningsEstimate_AP'] = td.CleanData(TD['Month12ChangeF12MEarningsEstimate_AP'],N) #Moderate outliers so model building is better
    TransformList = ['ZSBT','ZSCSBS'] #z-score by time, decile crosssection by sector
    NumTile = np.int_(5) #Number of quantiles to use in AnalyzeFactor
    TD['Month12ChangeF12MEarningsEstimate_AP'] = td.Transform(TD['Month12ChangeF12MEarningsEstimate_AP'],TransformList,D)
    FA['Month12ChangeF12MEarningsEstimate_AP'] = ft.AnalyzeFactor('Month12ChangeF12MEarningsEstimate_AP',TD['Month12ChangeF12MEarningsEstimate_AP'] ,TD['ForwardReturns_AP'],NumTile)
    FA.keys()
    FA['ICitp']
    FA['Price_Book_AP']
    #runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
    NumTile = np.int_(5)
    #runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')
    FA['Price_Book_AP'] = ft.AnalyzeFactor('Price_Book',TD['Price_Book_AP'],TD['ForwardReturns_AP'],NumTile)
    FA['ICitp'].keys()
    FA['Price_Book_AP'].keys()
    FA['ICitp']
    
    FA['Price_Book_AP']['ICitp']

    ([6.66104644e-01, 1.14284260e-01, 3.37005764e-02, 1.11534910e-01,
            1.89128057e-01, 1.43831658e-01, 4.68567515e-02, 1.69735981e-01,
            5.93813630e-02, 2.44547252e-03, 6.27103942e-02, 2.06584046e-01,
            1.18828626e-01, 4.27882483e-01, 3.28990942e-04, 5.73740851e-02,
            2.15897160e-03, 2.39335479e-07, 6.64520363e-07, 2.75015029e-04,
            1.04780149e-03, 6.16516523e-03, 4.72817573e-01, 1.78403944e-05,
            4.66891166e-03, 1.98299360e-02, 1.00655533e-02, 1.64708996e-03,
            8.94707052e-01, 5.00405931e-04, 1.36088675e-01, 1.85818313e-01,
            4.70780707e-04, 7.85616754e-02, 7.10839113e-04, 4.64021769e-03,
            7.06696447e-12, 1.51055508e-03, 6.91107530e-03, 1.02204658e-08,
            4.04312736e-07, 8.39886861e-01, 2.88862611e-03, 1.70519468e-10,
            9.05852259e-02, 3.17414287e-09, 9.69632027e-01, 6.37473323e-01,
            4.84851723e-01, 9.48558628e-02, 1.71344075e-01, 1.46611537e-02,
            1.75268564e-03, 7.96313018e-02, 4.85840322e-02, 2.30448159e-02,
            1.88948456e-02, 9.01748133e-01, 2.37058775e-01, 3.94954269e-01,
            1.27089036e-09, 2.79327741e-03, 9.61605271e-01, 9.05974908e-01,
            8.80837376e-01, 8.18793828e-07, 4.76067891e-12, 7.37301280e-02,
            6.92117806e-05, 2.38540757e-03, 7.49752358e-02, 3.95512422e-05,
            8.46139195e-01, 1.09104401e-05, 4.73318973e-03, 3.30693964e-01,
            1.76267928e-04, 2.09193493e-03, 6.67334940e-01, 1.92244817e-01,
            8.35959183e-01, 8.34191799e-01, 8.43540103e-02, 9.77876125e-01,
            2.21475455e-01, 1.88360202e-01, 9.81095400e-01, 3.21224926e-02,
            9.80688317e-04, 9.60680794e-01, 5.62860774e-02, 2.75247447e-02,
            5.28712513e-02, 1.09193134e-01, 9.45530460e-01, 2.23168226e-02,
            5.81498171e-01, 5.42233332e-01, 3.03734334e-02, 2.13373995e-01,
            1.47187712e-02, 6.18158002e-05, 2.56643909e-03, 2.46120427e-02,
            7.26602592e-02, 8.15872886e-01, 1.72858221e-01,            nan,
                       nan]),
      'GrandMeanIC': 0.04719828878543253,
      'GrandStdIC': 0.0719578551854404,
      'ICitp': 0.655915725445113,
      'ICdecay_n': array([0.04719829, 0.04604554, 0.04702866, 0.04815873, 0.048025  ,
            0.04823319, 0.04790066, 0.0454238 , 0.04477693, 0.04131434,
            0.04009083, 0.03740775]),
     'ICdecaySTD_n': array([0.07195786, 0.07083844, 0.06942435, 0.06785857, 0.06727422,
            0.06814571, 0.06789286, 0.06711164, 0.06562265, 0.06668908,
            0.06378868, 0.06311731]),
     'FactorIC_pn': array([[ 0.09234487, -0.01763312,  0.06086717, ...,  0.00328741,
              0.08467721,  0.08983206],
            [-0.0277634 ,  0.06893106,  0.07164694, ...,  0.08295416,
              0.0921836 ,  0.06816345],
            [ 0.05352306,  0.04581292,  0.08568681, ...,  0.09392823,
              0.10319894,  0.09828887],
            ...,
            [ 0.04342178,         nan,         nan, ...,         nan,
                     nan,         nan],
            [        nan,         nan,         nan, ...,         nan,
                     nan,         nan],
            [        nan,         nan,         nan, ...,         nan,
                     nan,         nan]]),
     'FactorICPval_pn': array([[0.00418864, 0.58529073, 0.05940401, ..., 0.91897554, 0.00866681,
             0.00534698],
            [0.39019575, 0.03271948, 0.02642826, ..., 0.01013102, 0.00425553,
             0.0347146 ],
            [0.09367729, 0.15141287, 0.00721661, ..., 0.00321721, 0.00120175,
             0.00204504],
            ...,
            [0.17285822,        nan,        nan, ...,        nan,        nan,
                    nan],
            [       nan,        nan,        nan, ...,        nan,        nan,
                    nan],
            [       nan,        nan,        nan, ...,        nan,        nan,
                    nan]]),
     'IRdecay_n': array([0.65591573, 0.65000779, 0.6774087 , 0.70969268, 0.71386937,
            0.70779504, 0.70553306, 0.67683932, 0.68233949, 0.61950686,
            0.62849445, 0.59267023]),
     'ICdecayWeighted': 0.046584500284270795,
     'AlphaBySignalTile_t': array([0.89177314, 0.94352203, 0.97853289, 0.9484604 , 0.93599462]),
     'BetaBySignalTile_t': array([ 0.01046579,  0.00439009, -0.00100789, -0.00468781, -0.00938737]),
     'RsquareBySignalTile_t': array([0.95195108, 0.97440474, 0.98164048, 0.9781223 , 0.96471854]),
     'TstatBySignalTile_t': array([1.31080114e-66, 1.54349942e-83, 1.56547249e-92, 8.79842855e-88,
            6.91190379e-75]),
     'ResidualRiskBySignalTile_t': array([0.02555823, 0.01939212, 0.01693873, 0.01797083, 0.02275679]),
     'ResidualBySignalTile_pt': array([[-0.87554018, -0.92869141, -0.96877924, -0.93211716, -0.92004414],
            [-0.90191345, -0.9477142 , -0.95521011, -0.92210952, -0.92248256],
            [-0.85822868, -0.92372311, -0.9446208 , -0.91262741, -0.91085851],
            [-0.87764741, -0.95486628, -0.95465579, -0.93010281, -0.91544818],
            [-0.94164879, -0.96617828, -1.00907361, -0.99158942, -0.955686  ],
            [-0.9901689 , -1.0323112 , -1.04962678, -1.00610852, -1.00831632],
            [-0.85543903, -0.90540467, -0.93760905, -0.9138893 , -0.8913225 ],
            [-0.85749447, -0.89895691, -0.94478791, -0.90989807, -0.89450557],
            [-0.92152476, -0.95225035, -0.98074831, -0.93055578, -0.9257596 ],
            [-0.87631696, -0.89765338, -0.92615728, -0.89673024, -0.8707292 ],
            [-0.91570358, -0.95438587, -0.98163996, -0.95875961, -0.93360828],
            [-0.88025934, -0.93764573, -0.95552168, -0.92006201, -0.8989775 ],
            [-0.93539107, -0.96030371, -0.98979757, -0.95106066, -0.93928225],
            [-0.95087894, -1.00377787, -1.02632375, -0.9826521 , -0.96731205],
            [-0.91120023, -0.93300087, -0.97018164, -0.93136052, -0.92792933],
            [-0.80780786, -0.88555183, -0.91213021, -0.87777841, -0.86894242],
            [-0.87131548, -0.90769088, -0.96375982, -0.9055295 , -0.90071498],
            [-0.82689239, -0.88919565, -0.9405846 , -0.89228716, -0.86134788],
            [-0.91840578, -0.9712975 , -0.9796411 , -0.95561429, -0.93348266],
            [-0.85424314, -0.89934123, -0.94379325, -0.88299227, -0.87760527],
            [-0.97597298, -0.99265351, -1.04273032, -0.9804398 , -0.96547095],
            [-0.87890747, -0.94322469, -0.98694552, -0.95914607, -0.9307438 ],
            [-0.89258003, -0.94414464, -0.97672888, -0.9449505 , -0.92207904],
            [-0.94046453, -0.95926236, -0.99121828, -0.96369566, -0.94218414],
            [-0.81828547, -0.89478778, -0.91354077, -0.89132478, -0.87654698],
            [-0.87943967, -0.91658483, -0.94289732, -0.93068452, -0.89462389],
            [-0.90743059, -0.9512604 , -0.99906882, -0.95473466, -0.93520078],
            [-0.95010078, -1.00344891, -1.03313327, -1.00128707, -0.98196775],
            [-0.94110065, -0.97329491, -0.99534573, -0.97457166, -0.95576585],
            [-0.99362769, -1.01234457, -1.05394055, -1.01659856, -1.0009106 ],
            [-1.11333027, -1.14443625, -1.16416525, -1.15509482, -1.10724875],
            [-0.86005345, -0.8921942 , -0.93452609, -0.88379599, -0.86305916],
            [-0.83851681, -0.87932874, -0.89984013, -0.88404183, -0.86471943],
            [-0.85274029, -0.89943701, -0.9375705 , -0.9147545 , -0.90529372],
            [-0.85549622, -0.90862702, -0.93302077, -0.91372681, -0.90550892],
            [-0.94522162, -0.97089543, -0.98162852, -0.97438467, -0.94073926],
            [-0.96213205, -1.00966005, -1.04696753, -1.01056584, -0.9743414 ],
            [-0.91318711, -0.9596318 , -0.9735141 , -0.93202251, -0.9056766 ],
            [-0.8628771 , -0.88966452, -0.90204548, -0.84612668, -0.82726067],
            [-0.91129278, -0.9662227 , -0.97376463, -0.93497846, -0.90356493],
            [-0.88794203, -0.93283764, -0.95926426, -0.91558703, -0.88444969],
            [-0.95044447, -0.98060335, -1.00781603, -0.98046413, -0.95347934],
            [-0.94350777, -0.99174152, -1.0267387 , -0.9866492 , -0.95671956],
            [-0.93238377, -0.97918317, -1.028679  , -0.98865286, -0.96320675],
            [-0.88046563, -0.91130316, -0.96719022, -0.96730929, -0.94874959],
            [-0.89740711, -0.95401571, -0.96628764, -0.94678928, -0.90669779],
            [-0.87774771, -0.94765299, -0.942375  , -0.91249783, -0.89533293],
            [-0.95696425, -1.01437055, -1.04188545, -0.98411456, -0.9749597 ],
            [-0.92378939, -0.96522748, -1.02686775, -0.9610251 , -0.94497113],
            [-0.84576192, -0.86947696, -0.89955652, -0.87746911, -0.85675573],
            [-0.92274865, -0.97750693, -0.97194305, -0.92334894, -0.92665421],
            [-0.92927267, -0.95239093, -0.97097969, -0.95491967, -0.93102197],
            [-0.92669243, -0.97251265, -0.99530005, -0.97370534, -0.94061355],
            [-0.92546383, -0.9635979 , -0.95336237, -0.92856532, -0.9034672 ],
            [-0.84773779, -0.89420497, -0.92895628, -0.89098976, -0.88322725],
            [-0.93437599, -0.95182263, -0.98872902, -0.94802419, -0.92860142],
            [-0.9158525 , -0.98539109, -0.99071883, -0.94265732, -0.91001162],
            [-1.00941845, -1.00814709, -1.01322881, -0.96881543, -0.95539815],
            [-0.8385244 , -0.86838351, -0.92103933, -0.85305128, -0.85503146],
            [-0.90816028, -0.9570427 , -0.94675514, -0.89786242, -0.86721014],
            [-0.976836  , -0.99349644, -0.99129916, -0.96527661, -0.93338617],
            [-0.97331943, -1.00313726, -1.00913291, -0.96753174, -0.97347726],
            [-0.83258751, -0.88852987, -0.90719975, -0.88639359, -0.84464428],
            [-0.89115353, -0.92902046, -0.95578571, -0.9209951 , -0.90199123],
            [-0.94493347, -0.98377934, -0.97357993, -0.94579466, -0.90772702],
            [-0.90901988, -0.97225416, -0.99599855, -0.95526826, -0.93622482],
            [-0.93760483, -0.99162365, -1.005331  , -0.9598811 , -0.9368423 ],
            [-1.00935107, -1.0920683 , -1.10699346, -1.08116474, -1.0685285 ],
            [-0.85341832, -0.913166  , -0.9511103 , -0.88209535, -0.88073074],
            [-0.8320425 , -0.88393086, -0.9092588 , -0.87425284, -0.86827552],
            [-0.86704652, -0.90948267, -0.93603345, -0.90330244, -0.89121163],
            [-0.90910596, -0.9382044 , -0.96430135, -0.95563504, -0.94352368],
            [-0.90035609, -0.93970431, -0.98542001, -0.95213838, -0.93020711],
            [-0.84946611, -0.88847943, -0.91126141, -0.88795988, -0.86270703],
            [-0.9118779 , -0.95959968, -0.98752565, -0.96743908, -0.93399712],
            [-0.93643086, -0.97490056, -1.01103662, -0.97536202, -0.95247904],
            [-0.98696107, -1.02798297, -1.0462603 , -1.0088373 , -1.0104796 ],
            [-0.9944766 , -1.05813546, -1.10437706, -1.05950456, -1.08125665],
            [-0.88371635, -0.94211913, -0.9888347 , -0.94660343, -0.92998955],
            [-0.9670638 , -1.05078496, -1.07593512, -1.05113089, -1.03402306],
            [-0.85505208, -0.92193315, -0.95057592, -0.90421306, -0.89257794],
            [-0.86116297, -0.8880452 , -0.91662824, -0.86904543, -0.82132563],
            [-0.93118705, -0.97867863, -1.01924461, -0.99915853, -0.9941547 ],
            [-0.92757733, -0.97404333, -1.01018299, -0.98954927, -0.97242494],
            [-0.91669977, -0.96252864, -1.01490573, -0.97856229, -0.95721665],
            [-0.87690465, -0.93354669, -0.98947825, -0.95769758, -0.90594595],
            [-0.82684319, -0.87345299, -0.90041048, -0.8585981 , -0.82572209],
            [-0.84570289, -0.86017287, -0.89814372, -0.85741307, -0.82458074],
            [-0.90159202, -0.93180522, -0.97573684, -0.93495523, -0.92859611],
            [-0.87601445, -0.92855673, -0.93382587, -0.90603924, -0.89176628],
            [-0.87857138, -0.91358386, -0.9293429 , -0.90000443, -0.89746858],
            [-0.89823118, -0.95569383, -1.00477968, -0.97242637, -0.94667994],
            [-0.83997713, -0.87851061, -0.91338991, -0.87512206, -0.8591909 ],
            [-0.87594114, -0.91852483, -0.95603283, -0.92097264, -0.90862421],
            [-0.86903518, -0.93530769, -0.94117671, -0.90875743, -0.8889292 ],
            [-0.88891549, -0.94388905, -0.9612823 , -0.92867982, -0.90703642],
            [-0.87835452, -0.92445886, -0.97428273, -0.91852753, -0.91001713],
            [-0.89827218, -0.94178281, -0.98566312, -0.95583423, -0.94688313],
            [-0.94014251, -0.97933629, -1.01431514, -0.96766807, -0.95585471],
            [-0.89011763, -0.93183405, -0.9630603 , -0.93499041, -0.91856537],
            [-0.88129375, -0.91624834, -0.95179098, -0.91896094, -0.91064984],
            [-0.92752038, -0.99535355, -1.02535578, -1.00062047, -0.96782934],
            [-0.91295086, -0.94155188, -0.98361394, -0.94252392, -0.93777618],
            [-0.86600384, -0.92216778, -0.94599217, -0.9289197 , -0.90782264],
            [-0.87694419, -0.92811816, -0.96030556, -0.93637243, -0.91940896],
            [-0.83744473, -0.88424639, -0.91556291, -0.89323462, -0.88522356],
            [-0.87459084, -0.91115849, -0.9467126 , -0.91628244, -0.89699168],
            [-0.93740232, -0.96916943, -1.00778325, -0.97736392, -0.96680837],
            [-0.86881762, -0.91292043, -0.95738311, -0.92583713, -0.92061149],
            [-0.92824705, -0.97398219, -1.00730292, -0.96899881, -0.94312434],
            [-0.92545492, -0.98440561, -1.01399896, -0.99268029, -0.97020084],
            [-0.85689779, -0.91350925, -0.94294253, -0.88859711, -0.90266531],
            [-0.88251376, -0.92275462, -0.96663246, -0.94520351, -0.92597586],
            [-0.84596186, -0.89971828, -0.92272258, -0.89847218, -0.89633273],
            [-0.92819871, -0.95601926, -0.99671987, -0.96829373, -0.95076383],
            [-0.89383148, -0.94538276, -0.97539023, -0.94185059, -0.93456447],
            [-0.93917058, -0.97710828, -1.01657467, -0.97638767, -0.94992083],
            [-0.85259598, -0.90043519, -0.94176912, -0.90819993, -0.89951412],
            [-0.90554544, -0.94226839, -0.98177622, -0.94909838, -0.94030337],
            [-0.83970405, -0.88545683, -0.93102393, -0.90891794, -0.89972372],
            [-0.89956105, -0.95746517, -0.98127263, -0.94538391, -0.93279023],
            [-0.86815396, -0.91573657, -0.96330536, -0.9346898 , -0.91930363],
            [-0.90731383, -0.94235814, -0.986457  , -0.94579135, -0.92176637],
            [-0.95704864, -0.99462139, -1.0068197 , -0.98562096, -0.9725297 ],
            [-0.88950073, -0.96408484, -1.00592238, -0.96021103, -0.94284052],
            [-0.9301474 , -0.97556459, -1.0003491 , -0.96478738, -0.94485112],
            [-0.87292393, -0.91811374, -0.96170715, -0.92490032, -0.9316823 ],
            [-0.88111285, -0.93028108, -0.96463572, -0.93163138, -0.90596345],
            [        nan,         nan,         nan,         nan,         nan]]),
     'ExcessRetBySignalTile_pt': array([[ 2.06735140e-03,  5.78033351e-04, -4.57620973e-03,
              1.96070031e-03,  1.50066222e-03],
            [-1.90329037e-02, -1.31393697e-02,  1.43270742e-02,
              1.73221057e-02,  4.44104778e-03],
            [ 4.58142409e-03, -9.34194669e-03,  4.61322905e-03,
              6.42642027e-03, -4.40801392e-03],
            [ 1.28104570e-03, -2.42678053e-02,  1.08834924e-02,
              5.31620904e-03,  7.44405167e-03],
            [-1.71955271e-02,  1.02245173e-02,  2.51833213e-03,
             -9.94843289e-03,  1.36444059e-02],
            [-2.15991977e-02, -1.15210877e-02,  6.59313049e-03,
              2.03244942e-02,  6.01563501e-03],
            [-2.04105820e-03, -4.93443767e-04,  2.10370702e-03,
             -4.39174403e-03,  5.52702204e-03],
            [-3.73554650e-03,  6.31748829e-03, -4.71001125e-03,
             -3.40321270e-05,  2.71214506e-03],
            [-2.72702258e-02, -6.23169889e-03,  2.94748993e-04,
              2.04240107e-02,  1.27661885e-02],
            [-2.92250104e-02,  9.13113825e-04,  7.17634716e-03,
              6.36473677e-03,  1.96877899e-02],
            [-1.49170393e-02, -1.79510011e-03,  6.01084504e-03,
             -1.14777658e-03,  1.15805725e-02],
            [-7.30520805e-03, -1.30582593e-02,  3.97400006e-03,
              9.29119610e-03,  1.78205406e-02],
            [-2.87250949e-02, -1.79741292e-03,  3.80083564e-03,
              1.25206453e-02,  1.19039928e-02],
            [-1.35232274e-02, -1.43934005e-02, -1.67975346e-03,
              1.20889245e-02,  1.51796638e-02],
            [-2.38652275e-02,  6.05576526e-03,  3.86165362e-03,
              1.25937788e-02,  3.53810383e-03],
            [ 1.61694588e-02, -1.02418952e-02, -2.17923520e-03,
              1.84795607e-03, -2.10381978e-03],
            [-8.85019306e-03,  6.34334993e-03, -1.48746052e-02,
              1.31742411e-02,  5.38378015e-03],
            [ 8.57100038e-03, -2.32911152e-03, -1.90143695e-02,
             -9.98825946e-04,  1.72072384e-02],
            [-1.50385761e-02, -1.61102196e-02,  1.06202921e-02,
              4.61772617e-03,  1.43386378e-02],
            [-1.00096688e-02, -3.65076711e-03, -1.33512436e-02,
              1.72004530e-02,  9.89588550e-03],
            [-3.27662240e-02,  2.61793238e-03, -1.21674367e-02,
              2.02418708e-02,  2.29892032e-02],
            [ 1.28425440e-02,  2.74061999e-04, -8.43602096e-03,
             -1.07091460e-02,  5.22722682e-03],
            [-4.71262746e-03, -4.55233679e-03, -2.14701317e-03,
             -4.55648359e-04,  9.93147634e-03],
            [-2.93830353e-02,  3.68656582e-03,  6.84684162e-03,
              4.36878316e-03,  1.35062171e-02],
            [ 1.31455096e-02, -1.19784064e-02,  3.95029564e-03,
             -4.13060203e-03, -2.10517530e-03],
            [-1.49290416e-02, -4.92703801e-04,  8.05695383e-03,
             -9.90412189e-03,  1.35612486e-02],
            [-5.82458314e-03,  2.15485911e-03, -1.05890522e-02,
              3.70918571e-03,  1.08239856e-02],
            [-4.43096409e-03, -5.69928033e-03, -7.87632909e-05,
              1.89538220e-03,  9.00488467e-03],
            [-2.19837279e-02, -2.26121035e-03,  1.08479939e-02,
              1.65126590e-03,  8.12115031e-03],
            [-2.64606959e-02,  7.03422602e-03,  8.60395051e-04,
              8.41026827e-03,  1.19905146e-02],
            [-2.65603164e-02, -4.72013782e-03,  1.16254562e-02,
             -8.65144585e-03,  2.76549341e-02],
            [-2.08761494e-02, -1.59095804e-03, -9.19889416e-03,
              1.12631266e-02,  1.92843751e-02],
            [-1.29448674e-02, -2.41438527e-03,  1.17239647e-02,
             -2.79642156e-03,  3.74578904e-03],
            [ 1.91784069e-03,  6.74212164e-03,  3.41703602e-03,
             -3.97748055e-03, -7.15875057e-03],
            [ 7.03207107e-04, -8.97129932e-04,  9.52592827e-03,
             -1.38488916e-03, -5.80173645e-03],
            [-3.06726140e-02, -4.45769603e-03,  1.99443207e-02,
             -2.79961174e-03,  1.84881802e-02],
            [-1.06949917e-02, -6.10777226e-03, -8.07891185e-03,
             -1.52783608e-03,  2.25141818e-02],
            [-2.40415084e-02, -1.87534516e-02,  2.36078505e-03,
              1.37701094e-02,  2.76377611e-02],
            [-4.38496591e-02, -1.93348462e-02,  2.89822469e-03,
              2.84740128e-02,  3.45287449e-02],
            [-2.03769608e-02, -2.35632513e-02,  3.90100527e-03,
              1.26115011e-02,  3.15551737e-02],
            [-1.95548215e-02, -1.28451350e-02, -4.38845865e-03,
              9.12931588e-03,  2.76897969e-02],
            [-2.41543064e-02, -2.35237600e-03,  5.63411254e-03,
              3.04188269e-03,  1.77248189e-02],
            [-1.05884591e-02, -6.82069385e-03, -6.58255135e-03,
              3.58747512e-03,  2.12467413e-02],
            [-2.50151115e-03,  2.68196126e-03, -1.15951176e-02,
             -1.49973959e-03,  1.16615775e-02],
            [ 6.36168290e-03,  2.72426741e-02,  6.33949549e-03,
             -2.38704577e-02, -1.78000290e-02],
            [-1.12298454e-02, -1.61239205e-02,  6.58449231e-03,
             -4.01045260e-03,  2.35886802e-02],
            [-1.07721801e-02, -2.90808362e-02,  1.10727540e-02,
              1.07852126e-02,  1.53665592e-02],
            [-1.09660262e-02, -1.62905054e-02, -8.49873512e-03,
              1.94013154e-02,  1.63479273e-02],
            [-7.08419864e-03,  3.37967602e-03, -2.31137201e-02,
              1.27491610e-02,  1.64557488e-02],
            [-2.27126643e-02,  4.89921926e-03,  9.45563178e-03,
              1.21497966e-03,  9.13619143e-03],
            [-2.59651301e-02, -2.89437651e-02,  1.16583128e-02,
              3.01985652e-02,  1.44513110e-02],
            [-2.96589228e-02, -9.80155692e-04,  1.54847266e-02,
              1.50141901e-03,  1.29705652e-02],
            [-1.36234490e-02, -7.56403406e-03,  4.77560636e-03,
             -3.62298072e-03,  1.71041598e-02],
            [-3.80344665e-02, -2.44463276e-02,  2.07763707e-02,
              1.54847859e-02,  2.80964866e-02],
            [-6.06404487e-03, -1.08995442e-03, -1.10368810e-03,
              6.60402569e-03,  1.66281989e-03],
            [-3.21623827e-02,  2.20396410e-03,  3.65401392e-04,
              1.10365704e-02,  1.80431324e-02],
            [-1.46594075e-02, -3.23912780e-02, -2.65676291e-03,
              1.53672909e-02,  3.55919405e-02],
            [-6.67678754e-02, -1.34352503e-02,  1.67714380e-02,
              3.13015382e-02,  3.24946583e-02],
            [-1.79462380e-02,  3.50639445e-03, -1.45269242e-02,
              2.31238650e-02,  8.33977989e-03],
            [-4.11438604e-02, -3.84293958e-02,  6.73398542e-03,
              2.54621443e-02,  4.35310668e-02],
            [-5.29801435e-02, -1.76947145e-02,  1.96884465e-02,
              1.57578208e-02,  3.53348388e-02],
            [-3.63656431e-02, -1.41571821e-02,  1.51044936e-02,
              2.68011947e-02,  8.60446461e-03],
            [-7.68942879e-03, -1.22935167e-02,  3.68267034e-03,
             -5.83236033e-03,  2.31335529e-02],
            [-1.88424777e-02, -5.08002127e-03,  3.05943094e-03,
              7.70517719e-03,  1.41508237e-02],
            [-4.18651741e-02, -2.88928029e-02,  1.63790876e-02,
              1.41338763e-02,  3.97893723e-02],
            [-3.20959803e-03, -1.46088102e-02, -3.26575437e-03,
              7.44424458e-03,  1.40885698e-02],
            [-1.90672591e-02, -2.11728651e-02,  2.76657805e-04,
              1.57535993e-02,  2.64537265e-02],
            [ 1.29345282e-02, -1.72324783e-02,  3.56519139e-03,
             -1.93227642e-04,  5.97071218e-04],
            [-4.84125734e-03, -1.31052806e-02, -1.62743432e-02,
              2.25074767e-02,  1.12011459e-02],
            [-5.58326883e-03, -6.12377070e-03,  3.20285959e-03,
              7.89344120e-03,  1.09478464e-03],
            [-1.31883974e-02, -4.10846090e-03,  4.14479774e-03,
              6.66231479e-03,  6.10727758e-03],
            [-1.50000334e-02,  7.66473425e-03,  1.65913812e-02,
             -4.80612501e-03, -5.14947312e-03],
            [-6.72498861e-03,  5.68708384e-03, -5.00762017e-03,
             -1.79156561e-03,  7.68274441e-03],
            [-1.68523479e-02, -4.48001606e-03,  7.42614922e-03,
              4.35188670e-04,  1.29412910e-02],
            [-7.54702833e-03, -3.44282461e-03,  3.71057880e-03,
             -6.22864461e-03,  1.48071722e-02],
            [-1.51162673e-02, -1.65569678e-03, -2.61974400e-03,
              3.09223209e-03,  1.36497198e-02],
            [-1.91383002e-02, -7.94436926e-03,  9.20402778e-03,
              1.68373474e-02,  3.09044414e-03],
            [ 1.56528066e-02,  4.46953480e-03, -6.11555659e-03,
              9.12460225e-03, -2.45311734e-02],
            [ 6.05660147e-03, -6.09580686e-04, -1.23251876e-02,
             -1.73843860e-04,  3.96474608e-03],
            [ 1.98486654e-02, -1.15394561e-02, -1.15975918e-03,
             -6.07421387e-03, -9.80321122e-04],
            [ 2.31058229e-03, -1.30328811e-02, -6.85249185e-03,
              9.30991579e-03,  8.31581711e-03],
            [-3.70413984e-02, -1.25901166e-02, -6.53133982e-03,
              1.07273948e-02,  4.56601183e-02],
            [ 5.03745447e-03,  9.56768122e-03,  4.25506083e-03,
             -5.56604915e-03, -1.28168919e-02],
            [-1.32295409e-03,  4.17164150e-03,  3.23095691e-03,
             -6.07958506e-03, -1.25728616e-03],
            [ 1.06003177e-03,  7.13960961e-03, -1.00848603e-02,
             -3.71726575e-03,  5.28600408e-03],
            [ 7.99023156e-03,  3.05484199e-03, -1.79033696e-02,
             -1.62207664e-02,  2.30324109e-02],
            [-1.66922874e-02, -1.20543525e-02, -4.44623793e-03,
              6.99011344e-03,  2.70126975e-02],
            [-3.52571353e-02,  1.52241638e-03, -1.88120554e-03,
              8.47450416e-03,  2.84548009e-02],
            [-1.45710375e-02,  6.93545949e-03, -2.01121170e-03,
              8.68023825e-03,  2.55099641e-03],
            [-1.62394914e-02, -1.72293570e-02,  1.23378206e-02,
              9.93296863e-03,  1.15881647e-02],
            [-2.22429862e-02, -5.72421598e-03,  1.33342548e-02,
              1.24684255e-02,  2.37014925e-03],
            [ 9.23816591e-03,  3.62077203e-03, -1.03685817e-02,
             -8.02939283e-03,  5.32579933e-03],
            [-1.39426492e-02, -1.13088004e-03, -1.35791871e-03,
              6.59297519e-03,  9.74613517e-03],
            [-7.89136778e-03,  1.12816284e-03, -1.49838206e-03,
              3.40109434e-03,  3.17107295e-03],
            [-8.10375521e-03, -2.28167515e-02,  6.15685730e-03,
              8.38894604e-03,  1.56049162e-02],
            [-9.94741414e-03, -1.32507090e-02,  4.29705744e-03,
              6.77943201e-03,  1.58962267e-02],
            [-4.69745490e-03,  8.35855119e-04, -1.40759719e-02,
              1.15393679e-02,  7.49794442e-03],
            [-4.74529914e-04,  7.80071644e-03, -1.03586358e-03,
             -1.25706271e-03, -4.74312987e-03],
            [-1.70515091e-02, -4.30412168e-03, -4.10125643e-03,
              1.25897868e-02,  1.20860955e-02],
            [-9.92750994e-03,  3.38297464e-05,  3.75525823e-03,
              1.70958670e-03,  5.61383007e-03],
            [-1.31371267e-02,  3.51216352e-03,  2.85156676e-03,
              5.52129172e-03,  1.25443870e-03],
            [ 7.40243838e-03, -8.41691667e-03, -3.17288802e-03,
             -8.34960519e-03,  1.21806716e-02],
            [-1.74341133e-02,  5.73673577e-03, -1.29403207e-03,
              9.73741708e-03,  2.03715438e-03],
            [ 6.70618070e-04, -3.89853951e-03,  7.15101968e-03,
             -5.94233935e-03,  2.56973956e-03],
            [-3.51517141e-04,  1.30189146e-04,  2.87084756e-03,
             -3.32497210e-03,  1.10061636e-03],
            [-6.45580649e-04,  3.96410555e-03,  7.35856104e-03,
             -5.90073653e-04, -5.30588875e-03],
            [-1.32809182e-02,  1.71327367e-03,  1.00384922e-03,
              1.24822773e-03,  7.92852327e-03],
            [-1.38476073e-02,  6.32930298e-03,  2.89972124e-03,
              3.36475262e-03,  1.60545265e-03],
            [ 8.51496836e-04,  8.36184988e-03, -1.21053449e-03,
              1.80746876e-04, -7.16437525e-03],
            [-1.23207725e-02, -6.15872878e-03, -4.33683354e-03,
              3.98460896e-03,  1.75079998e-02],
            [ 3.51658339e-03, -3.45683266e-03,  2.16360962e-03,
             -6.45188125e-03,  3.73845223e-03],
            [-3.37599082e-03, -8.47343772e-03, -3.10450659e-03,
              2.10261741e-02, -5.68947644e-03],
            [-1.51960956e-03,  9.92222829e-03,  9.96456462e-04,
             -7.68716892e-03, -9.76501988e-04],
            [ 2.23787737e-04, -2.06365207e-03,  9.69423937e-03,
              3.70261634e-03, -6.84023426e-03],
            [-1.63035414e-02,  7.74834624e-03,  2.16837612e-03,
              5.96848314e-04,  5.75652988e-03],
            [-3.40995252e-03, -3.22064787e-03,  1.77537635e-03,
              5.23749547e-03,  5.14077141e-05],
            [-1.55417076e-02, -1.53492490e-03, -5.81667621e-03,
              4.41630406e-03,  1.85686448e-02],
            [ 4.12838718e-04,  4.08450117e-03, -2.45002825e-03,
              9.02514260e-04, -3.06156235e-03],
            [-1.00687925e-02,  4.97987268e-03,  5.03114065e-04,
              3.12224109e-03, -5.30944533e-04],
            [ 6.01935761e-03,  1.17327221e-02,  9.25286850e-04,
             -7.21246170e-03, -1.07027341e-02],
            [-4.37764464e-03, -1.05119328e-02,  7.10071241e-04,
              6.53898226e-03,  6.68307314e-03],
            [ 4.29999838e-03,  8.34765545e-03, -4.31565825e-03,
             -5.84442888e-03, -3.01579890e-03],
            [-1.49389339e-02,  1.76933848e-03, -7.31536632e-03,
              3.28002498e-03,  1.48420827e-02],
            [-2.25738928e-02, -8.13557160e-03,  1.49099282e-02,
              6.19497542e-03,  7.02325121e-03],
            [ 1.49633984e-02, -7.79390339e-03, -1.45513411e-02,
              1.13470961e-03,  6.09970489e-03],
            [-1.54333827e-02, -8.96083454e-03,  1.39067302e-03,
              6.96521088e-03,  1.45446413e-02],
            [ 9.74048565e-04,  7.42337366e-03, -1.25667922e-03,
              5.41118809e-03, -1.39214778e-02],
            [-6.26158937e-03, -3.78482851e-03, -3.22091292e-03,
             -3.51989066e-04,  1.27697863e-02],
            [            nan,             nan,             nan,
                         nan,             nan]]),
     'CumulativeTileReturn_t': array([0.22354213, 0.55393623, 1.14100093, 1.89111285, 3.60766733]),
     'CumulativeTileExcessReturn_t': array([0.22354213, 0.55393623, 1.14100093, 1.89111285, 3.60766733]),
     'ExPostIR_t': array([120.86881609, 168.5455926 , 200.11757721, 182.82753562,
            142.47971766])} 
runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')

NumTile                 = np.int_(5)

runfile('/Users/kriti/Downloads/Deepak/Midterm2/Example_Combine_Factors.py', wdir='/Users/kriti/Downloads/Deepak/Midterm2')

FA['Price_Book_AP']     = ft.AnalyzeFactor('Price_Book',TD['Price_Book_AP'],TD['ForwardReturns_AP'],NumTile)
#/Users/kriti/Downloads/Deepak/Midterm2/FactorTester.py:151: RuntimeWarning: Mean of empty slice
R['EW_BenchMark_p']                 =   np.nanmean(Returns_AP,axis=0)#equaly weighted mean by period

FA['ICitp'].keys()
    
